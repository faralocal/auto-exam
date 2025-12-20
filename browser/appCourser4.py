#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Playwright workflow runner (sync, Chromium persistent profile)

- Reads a JSON workflow (list of steps) and executes them in order.
- Supported step types:
    - "goto": open a URL (value or url key)
    - "click": find an element (by tag/attr/value/class/text) and click it
               (supports array_select_one to pick index when multiple)
    - "array": find multiple parent elements (by tag/class/attr/value),
               optionally filter by inner text (if_find_text_inside),
               then within each parent click child matchers listed in "click" array
    - "frame": switch to an iframe (by selector, name, or URL)
    - "main_frame": switch back to the main frame
    - "condition": execute steps based on conditions
    - "write": type text with random delays
    - "use_last_tab": switch to the last opened tab
    - "scroll": scroll to element or position
    - "download_from_link": click a link and save the downloaded file
    - "download_page": save the current page as HTML or plain text
    - "group_action": find multiple elements and run nested actions on each (can be nested)
- All logs are in English and saved to workflow.log. On any failure the run stops.
- Tolerant to minor key typos like "Title" and "arrt".
"""

import argparse
import ctypes
import json
import logging
import os
import random
import re
import sys
import time
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urljoin, urlparse

from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PWTimeout
from playwright.sync_api import sync_playwright


LOG_CAPTURE_LIST = []

def load_excel_rows(file_path: str, start_row: int = 2) -> List[List[str]]:
    """
    Load rows from an Excel (.xlsx) file starting from `start_row` (1-based).
    Stops scanning when it reaches the first fully-empty row (end-of-data marker).
    Returns a list of rows, each row is a list of cell values (as strings).
    """
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows: List[List[str]] = []
    started = False

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx < start_row:
            continue

        started = True

        # Convert all cells to string (None → "")
        clean_row = [str(cell) if cell is not None else "" for cell in row]

        # Stop at the first fully-empty row
        # (treat whitespace-only values as empty too)
        if all((c.strip() == "") for c in clean_row):
            logger.info(
                f"🛑 Reached an empty Excel row at index {idx}. Stopping Excel scan."
            )
            break

        rows.append(clean_row)

    wb.close()

    if started:
        logger.info(f"📊 Loaded {len(rows)} data rows from Excel (starting at row {start_row})")
    else:
        logger.info("📊 Excel scan did not start (start_row beyond sheet range).")

    return rows



def exec_step_write_excel(
    page, step: Dict[str, Any], current_row: List[str], current_frame=None, parent=None
) -> None:
    """
    Write a value from the current Excel row into a text field.
    - `write_from_col`: 1-based column index (e.g., 1 = first column)
    """
    col_index = to_int_or_none(get_key(step, "write_from_col"))
    if col_index is None:
        raise RuntimeError('write_excel requires "write_from_col" (1-based index).')
    if col_index < 1:
        raise RuntimeError('"write_from_col" must be >= 1.')

    # Get value from row (0-based internally)
    cell_value = ""
    if col_index - 1 < len(current_row):
        cell_value = current_row[col_index - 1]
    else:
        logger.warning(
            f"⚠️ Column {col_index} not found in row (row has {len(current_row)} columns). Using empty string."
        )

    # Now reuse exec_step_write logic, but with `cell_value` as text
    tag = get_key(step, "tag")
    attr = get_key(step, "attr", "arrt", "attribute")
    value = get_key(step, "value")
    cls = get_key(step, "class")
    text_filter = get_key(step, "text")
    idx = to_int_or_none(get_key(step, "array_select_one"))
    ignore_error = get_key(step, "ignore", default=False)
    timeout = float(get_key(step, "timeout", default=35000))

    selector = build_css_selector(tag, cls, attr, value)
    root = get_locator_root(page, current_frame, parent)
    loc = root.locator(selector)
    if text_filter:
        loc = loc.filter(has_text=text_filter)

    logger.info(
        f"⌨️ [Excel] Writing '{cell_value}' (from col {col_index}) to: {selector}"
    )

    try:
        if idx is None:
            idx = 0
        count = loc.count()
        if count == 0:
            if ignore_error:
                logger.warning(
                    f"⚠️ No elements found for write_excel but ignoring: {selector}"
                )
                return
            else:
                raise RuntimeError(f"No elements found for write_excel: {selector}")
        if idx < 0 or idx >= count:
            if ignore_error:
                logger.warning(f"⚠️ Index {idx} out of range (found {count}), ignoring.")
                return
            else:
                raise RuntimeError(f"Index {idx} out of range (found {count}).")

        target = loc.nth(idx)
        target.wait_for(state="visible", timeout=timeout)
        target.scroll_into_view_if_needed()
        target.click()
        if get_key(step, "clear", default=True):
            target.clear()
        human_type(target, cell_value)
    except Exception as e:
        if ignore_error:
            logger.warning(f"⚠️ write_excel failed but ignoring: {e}")
        else:
            raise

    step_sleep(get_key(step, "sleep"))


# ------------------ Logging ------------------
LOG_FILE = "workflow.log"
logger = logging.getLogger("workflow")
logger.setLevel(logging.INFO)
fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
ch = logging.StreamHandler()
fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
fh.setFormatter(fmt)
ch.setFormatter(fmt)
logger.addHandler(fh)
logger.addHandler(ch)


import os
import time

import requests


def exec_step_group_excel(
    page, browser, step: Dict[str, Any], current_frame=None, parent=None
) -> None:
    """
    group_excel:
    - Reads an Excel file (.xlsx)
    - Starts from `start_row` (default: 2)
    - Stops when it reaches the first fully-empty row
    - For each row, runs `actions` with access to row data via context
    Supports in actions:
      - "write_excel": uses `write_from_col` (1-based index) to get value from current row
    """
    file_path = get_key(step, "file")
    start_row = to_int_or_none(get_key(step, "start_row")) or 2
    actions: List[Dict[str, Any]] = get_key(step, "actions", "steps", default=[])
    ignore_error = get_key(step, "ignore", default=False)

    if not file_path:
        raise RuntimeError('group_excel requires "file" key.')
    if not actions:
        raise RuntimeError('group_excel requires non-empty "actions" array.')

    rows = load_excel_rows(file_path, start_row=start_row)
    if not rows:
        logger.warning("⚠️ Excel file has no data rows (after start_row). Skipping actions.")
        return

    logger.info(f"🧮 Processing {len(rows)} Excel rows...")

    for row_index, current_row in enumerate(rows):
        excel_row_number = row_index + start_row

        # Extra safety: if a blank row slips in, stop immediately
        if all((str(c).strip() == "") for c in current_row):
            logger.info(f"🛑 Empty Excel row detected at {excel_row_number}. Stopping iteration.")
            break

        logger.info(f"🧮 [Excel Row {excel_row_number}] Processing...")
        local_frame = current_frame

        for j, action in enumerate(actions, start=1):
            a_title = get_key(action, "title", "Title", default=f"Excel action #{j}")
            a_type = get_key(action, "type")
            if not a_type:
                logger.warning("⚠️ [group_excel] Missing 'type' in action, skipping.")
                continue

            stype_l = str(a_type).strip().lower()
            logger.info(f"   ▶️ [Excel Row {excel_row_number}] Action {j}: {a_title} ({stype_l})")

            action_ignore = get_key(action, "ignore", default=False)

            try:
                if stype_l == "write_excel":
                    exec_step_write_excel(
                        page,
                        action,
                        current_row,
                        current_frame=local_frame,
                        parent=parent,
                    )
                elif stype_l == "click":
                    exec_step_click(page, action, local_frame, parent=parent)
                elif stype_l == "write":
                    exec_step_write(page, action, local_frame, parent=parent)
                elif stype_l == "scroll":
                    exec_step_scroll(page, action, local_frame, parent=parent)
                elif stype_l == "array":
                    exec_step_array(page, action, local_frame, parent=parent)
                elif stype_l == "group_action":
                    exec_step_group_action(page, browser, action, local_frame, parent=parent)
                elif stype_l == "download_from_link":
                    exec_step_download_from_link(page, action, local_frame, parent=parent)
                elif stype_l == "use_last_tab":
                    exec_step_use_last_tab(browser, action)
                elif stype_l == "goto":
                    exec_step_goto(page, action)
                    local_frame = None
                elif stype_l == "frame":
                    local_frame = exec_step_frame(page, action)
                elif stype_l == "main_frame":
                    local_frame = exec_step_main_frame(page, action)
                elif stype_l == "refresh":
                    exec_step_refresh(page, action)
                elif stype_l == "select":
                    exec_step_select(page, action, local_frame, parent=parent)
                else:
                    if action_ignore or ignore_error:
                        logger.warning(f"⚠️ Unsupported action type in group_excel but ignoring: '{a_type}'")
                    else:
                        raise RuntimeError(f"[group_excel] Unsupported action type: '{a_type}'")
            except Exception as e:
                if action_ignore or ignore_error:
                    logger.warning(f"⚠️ [group_excel] Action failed but ignoring: {a_title} | {e}")
                    continue
                else:
                    raise

    step_sleep(get_key(step, "sleep"))


# ------------------ Desktop size detection ------------------
def get_desktop_size() -> Tuple[int, int]:
    """Cross-platform best-effort screen size detection."""
    try:
        user32 = ctypes.windll.user32 if hasattr(ctypes, "windll") else None
        if user32:
            return user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)
        else:
            import subprocess

            wh = subprocess.check_output(
                "xrandr | grep '*' | awk '{print $1}'", shell=True
            )
            w, h = map(int, wh.decode().strip().split("x"))
            return w, h
    except Exception:
        # Fallback
        return 1366, 768


# ------------------ Human typing (optional utility) ------------------
def human_type(element, text: str):
    """Type like a human: small random delays; slow down on spaces."""
    for ch in text:
        element.type(ch)
        extra = random.randint(100, 200) / 1000 if ch == " " else 0
        time.sleep(random.randint(50, 150) / 1000 + extra)


# ------------------ Helpers ------------------
def get_key(d: Dict[str, Any], key: str, *alts: str, default=None):
    """Fetch d[key] with tolerant aliasing (e.g., attr/arrt/attribute)."""
    if key in d:
        return d[key]
    for a in alts:
        if a in d:
            return d[a]
    # Fix common case-insensitive
    for k in d.keys():
        if k.lower() == key.lower():
            return d[k]
    return default


def to_int_or_none(x) -> Optional[int]:
    if x is None:
        return None
    try:
        return int(x)
    except Exception:
        return None


def normalize_class_selector(cls_value: Optional[str]) -> str:
    """Return CSS class part like '.c1.c2' or '' if none."""
    if not cls_value:
        return ""
    s = cls_value.strip()
    if s.startswith("."):
        # could be ".c1.c2" already
        return s
    # allow space-separated classes
    parts = [p for p in s.split() if p]
    return "." + ".".join(parts) if parts else ""


def build_css_selector(
    tag: Optional[str],
    cls: Optional[str],
    attr: Optional[str],
    value: Optional[str],
) -> str:
    """Build a robust CSS selector from parts."""
    t = (tag or "*").strip()
    c = normalize_class_selector(cls)
    a = ""
    if attr and value is not None:
        a = f'[{attr}="{value}"]'
    elif attr:
        a = f"[{attr}]"
    return f"{t}{c}{a}"


def wait_and_click(
    loc, index: int = 0, timeout: float = 35000, ignore_error: bool = False
):
    try:
        count = loc.count()
        if count == 0:
            if ignore_error:
                logger.warning("🚫 No matching elements found, but ignoring error.")
                return False
            else:
                raise RuntimeError("🚫 No matching elements found.")

        if index < 0 or index >= count:
            if ignore_error:
                logger.warning(
                    f"🚫 array_select_one index {index} is out of range (found {count}), but ignoring error."
                )
                return False
            else:
                raise RuntimeError(
                    f"array_select_one index {index} is out of range (found {count})."
                )

        target = loc.nth(index)
        target.wait_for(state="visible", timeout=timeout)
        target.scroll_into_view_if_needed()

        # ذخیره وضعیت قبل از کلیک (آیا المان href دارد؟)
        is_link = False
        try:
            is_link = bool(target.get_attribute("href"))
        except Exception:
            pass  # نادیده گرفتن خطا در صورت منقضی بودن المان

        # اجرای کلیک
        target.click(timeout=timeout)

        # اگر المان لینک بود، منتظر ناوبری شویم
        if is_link:
            try:
                page = target.page
                page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                time.sleep(2)  # فول‌بک در صورت خطا
        return True

    except Exception as e:
        if ignore_error:
            logger.warning(f"⚠️ Click failed but ignoring: {str(e).split(':')[0]}")
            return False
        else:
            raise RuntimeError(
                f"Element interaction failed: {str(e).split(':')[0]}"
            ) from e


def step_sleep(seconds: Optional[float]):
    if seconds is None:
        return
    try:
        s = float(seconds)
    except Exception:
        s = 0
    if s > 0:
        time.sleep(s)


def make_safe_filename(name: str, default: str, ext: str) -> str:
    """Sanitize a filename (very simple) and ensure extension."""
    base = (name or "").strip() or default
    # Remove characters that are problematic in filenames
    base = re.sub(r'[\\/*?:"<>|]', "_", base)
    if ext and not base.lower().endswith(ext.lower()):
        base += ext
    return base


def get_locator_root(page, current_frame=None, parent=None):
    """
    Decide which object to use as root for .locator() calls.
    Priority:
      1) parent (Locator from group_action)
      2) current_frame (Frame or FrameLocator)
      3) page
    """
    if parent is not None:
        return parent
    if current_frame is not None:
        return current_frame
    return page


# ------------------ Condition Checking ------------------
def check_condition(
    page, condition: Dict[str, Any], current_frame=None, parent=None
) -> bool:
    """
    Check a condition based on element presence/absence.
    Supported condition types:
    - "status": "found" or "not_found"
    - "tag", "attr", "value", "class", "text": element selector parameters
    """
    status = get_key(condition, "status")
    tag = get_key(condition, "tag")
    attr = get_key(condition, "attr", "arrt", "attribute")
    value = get_key(condition, "value")
    cls = get_key(condition, "class")
    text = get_key(condition, "text")

    if not status:
        raise RuntimeError('Condition missing "status" (found/not_found)')

    selector = build_css_selector(tag, cls, attr, value)

    root = get_locator_root(page, current_frame, parent)
    loc = root.locator(selector)

    if text:
        loc = loc.filter(has_text=text)

    count = loc.count()

    logger.info(
        f"🔍 Condition check: {selector} status={status}, found={count} elements"
    )

    if status == "found":
        return count > 0
    elif status == "not_found":
        return count == 0
    else:
        raise RuntimeError(f'Unknown condition status: "{status}"')


# ------------------ Frame Management ------------------
def switch_to_frame(page, step: Dict[str, Any]):
    """
    Switch to an iframe based on selector, name, or URL.
    Supports:
    - "selector": CSS selector for the iframe
    - "name": name attribute of the iframe
    - "url": URL of the iframe (or partial match)
    - "index": numerical index of the iframe
    """
    frame_selector = get_key(step, "selector")
    frame_name = get_key(step, "name")
    frame_url = get_key(step, "url")
    frame_index = to_int_or_none(get_key(step, "index"))

    if frame_selector:
        logger.info(f"🖼️ Switching to frame by selector: {frame_selector}")
        frame = page.frame_locator(frame_selector)
        return frame
    elif frame_name:
        logger.info(f"🖼️ Switching to frame by name: {frame_name}")
        frame = page.frame(name=frame_name)
        if not frame:
            raise RuntimeError(f"Frame with name '{frame_name}' not found.")
        return frame
    elif frame_url:
        logger.info(f"🖼️ Switching to frame by URL: {frame_url}")
        for frame in page.frames:
            if frame_url in frame.url:
                return frame
        raise RuntimeError(f"Frame with URL containing '{frame_url}' not found.")
    elif frame_index is not None:
        logger.info(f"🖼️ Switching to frame by index: {frame_index}")
        frames = page.frames
        if frame_index < 0 or frame_index >= len(frames):
            raise RuntimeError(
                f"Frame index {frame_index} out of range (0-{len(frames) - 1})"
            )
        return frames[frame_index]
    else:
        raise RuntimeError(
            'Frame step requires one of: "selector", "name", "url", or "index"'
        )


def switch_to_main_frame(page):
    """Switch back to the main frame."""
    logger.info("🏠 Switching back to main frame")
    # In Playwright, we're automatically in the main frame when we don't specify a frame
    return None


def exec_step_select(
    page, step: Dict[str, Any], current_frame=None, parent=None
) -> None:
    """
    Select an option in a <select> element.
    Supported keys:
      - tag, class, attr, value, text: to locate the <select>
      - option_value: value of the <option> to select (optional)
      - option_label: visible text of the <option> to select (optional)
      - option_index: index of the <option> to select (optional)
      - array_select_one: if multiple <select> elements match, which one to use (default: 0)
    At least one of option_value, option_label, or option_index must be provided.
    """
    tag = get_key(step, "tag", default="select")
    attr = get_key(step, "attr", "arrt", "attribute")
    value = get_key(step, "value")
    cls = get_key(step, "class")
    idx = to_int_or_none(get_key(step, "array_select_one"))
    ignore_error = get_key(step, "ignore", default=False)

    # Option selection criteria
    option_value = get_key(step, "option_value")
    option_label = get_key(step, "option_label")
    option_index = to_int_or_none(get_key(step, "option_index"))

    if not any(
        [option_value is not None, option_label is not None, option_index is not None]
    ):
        raise RuntimeError(
            'select step requires one of: "option_value", "option_label", or "option_index"'
        )

    selector = build_css_selector(tag, cls, attr, value)
    root = get_locator_root(page, current_frame, parent)
    loc = root.locator(selector)

    logger.info(f"📋 Select selector: {selector}")
    try:
        if idx is None:
            idx = 0
        count = loc.count()
        if count == 0:
            if ignore_error:
                logger.warning(f"⚠️ No <select> found but ignoring: {selector}")
                return
            else:
                raise RuntimeError(f"No <select> element found: {selector}")
        if idx < 0 or idx >= count:
            if ignore_error:
                logger.warning(
                    f"⚠️ array_select_one index {idx} out of range (found {count}), ignoring."
                )
                return
            else:
                raise RuntimeError(
                    f"array_select_one index {idx} out of range (found {count})."
                )

        target_select = loc.nth(idx)
        target_select.wait_for(
            state="visible", timeout=float(get_key(step, "timeout", default=35000))
        )
        target_select.scroll_into_view_if_needed()

        # Build selection args for select_option()
        select_args = {}
        if option_value is not None:
            select_args["value"] = option_value
        if option_label is not None:
            select_args["label"] = option_label
        if option_index is not None:
            select_args["index"] = option_index

        logger.info(f"  → Selecting option: {select_args}")
        target_select.select_option(**select_args)

    except Exception as e:
        if ignore_error:
            logger.warning(f"⚠️ Select failed but ignoring: {e}")
        else:
            raise RuntimeError(f"Select step failed: {e}") from e

    step_sleep(get_key(step, "sleep"))


# ------------------ Step executors ------------------
def exec_step_goto(page, step: Dict[str, Any]) -> None:
    url = get_key(step, "value", "url")
    if not url:
        raise RuntimeError('Missing "value" or "url" for goto step.')
    logger.info(f"🌐 Navigating to: {url}")
    page.goto(url)
    step_sleep(get_key(step, "sleep"))


def exec_step_click(
    page, step: Dict[str, Any], current_frame=None, parent=None
) -> None:
    # Check condition first
    condition = get_key(step, "if")
    if condition:
        condition_met = check_condition(page, condition, current_frame, parent)
        logger.info(f"🔍 Condition check result: {condition_met}")

        if condition_met:
            # Execute alternative click steps
            alt_clicks = get_key(condition, "click", default=[])
            if not isinstance(alt_clicks, list):
                alt_clicks = [alt_clicks]

            for alt_click in alt_clicks:
                if not isinstance(alt_click, dict):
                    continue

                logger.info("🔄 Executing alternative click due to condition")
                # Recursively execute click step with alternative configuration
                exec_step_click(page, alt_click, current_frame, parent)
            return  # Don't execute main click if condition was met and alternative executed

    # Proceed with normal click execution if no condition or condition not met
    tag = get_key(step, "tag")
    attr = get_key(step, "attr", "arrt", "attribute")
    value = get_key(step, "value")
    cls = get_key(step, "class")
    text = get_key(step, "text")
    idx = to_int_or_none(get_key(step, "array_select_one"))
    ignore_error = get_key(step, "ignore", default=False)

    selector = build_css_selector(tag, cls, attr, value)

    root = get_locator_root(page, current_frame, parent)
    loc = root.locator(selector)

    if text:
        loc = loc.filter(has_text=text)

    logger.info(f"🔘 Click selector: {selector}{' | has_text=' + text if text else ''}")
    try:
        if idx is None:
            idx = 0
        success = wait_and_click(
            loc,
            index=idx,
            timeout=float(get_key(step, "timeout", default=45000)),
            ignore_error=ignore_error,
        )
        if not success and ignore_error:
            return
    except PWTimeout as e:
        if ignore_error:
            logger.warning(f"⚠️ Timeout waiting for element but ignoring: {selector}")
            return
        else:
            raise RuntimeError(f"Timeout waiting for element: {selector}") from e

    step_sleep(get_key(step, "sleep"))


def exec_step_write(
    page, step: Dict[str, Any], current_frame=None, parent=None
) -> None:
    """Type text with human-like delays."""
    text = get_key(step, "write", "value", "text")
    if not text:
        raise RuntimeError('Missing "write" or "value" for write step.')

    tag = get_key(step, "tag")
    attr = get_key(step, "attr", "arrt", "attribute")
    value = get_key(step, "value")
    cls = get_key(step, "class")
    text_filter = get_key(step, "text")
    idx = to_int_or_none(get_key(step, "array_select_one"))
    ignore_error = get_key(step, "ignore", default=False)

    selector = build_css_selector(tag, cls, attr, value)

    root = get_locator_root(page, current_frame, parent)
    loc = root.locator(selector)

    if text_filter:
        loc = loc.filter(has_text=text_filter)

    logger.info(f"⌨️ Writing '{text}' to selector: {selector}")

    try:
        if idx is None:
            idx = 0

        count = loc.count()
        if count == 0:
            if ignore_error:
                logger.warning(
                    f"⚠️ No elements found for writing but ignoring: {selector}"
                )
                return
            else:
                raise RuntimeError(f"No elements found for writing: {selector}")

        if idx < 0 or idx >= count:
            if ignore_error:
                logger.warning(
                    f"⚠️ array_select_one index {idx} is out of range (found {count}), but ignoring error."
                )
                return
            else:
                raise RuntimeError(
                    f"array_select_one index {idx} is out of range (found {count})."
                )

        target = loc.nth(idx)
        target.wait_for(
            state="visible", timeout=float(get_key(step, "timeout", default=35000))
        )

        # Scroll to element
        target.scroll_into_view_if_needed()

        # Click to focus and clear if needed
        target.click()
        if get_key(step, "clear", default=True):
            target.clear()

        # Type with human-like delays
        human_type(target, text)

    except Exception as e:
        if ignore_error:
            logger.warning(f"⚠️ Write failed but ignoring: {e}")
        else:
            raise


def exec_step_array(
    page, step: Dict[str, Any], current_frame=None, parent=None
) -> None:
    """
    Find multiple parent elements by tag/class/attr/value,
    optionally filter by inner text (if_find_text_inside),
    then for each (or selected one) click child matchers defined in 'click' list.
    """
    tag = get_key(step, "tag")
    attr = get_key(step, "attr", "arrt", "attribute")
    value = get_key(step, "value")
    cls = get_key(step, "class")
    filter_text = get_key(step, "if_find_text_inside")
    parent_idx = to_int_or_none(get_key(step, "array_select_one"))  # optional
    ignore_error = get_key(step, "ignore", default=False)

    parent_selector = build_css_selector(tag, cls, attr, value)

    root = get_locator_root(page, current_frame, parent)
    parents = root.locator(parent_selector)

    if filter_text:
        parents = parents.filter(has_text=filter_text)

    total = parents.count()
    if total == 0:
        if ignore_error:
            logger.warning(
                f"⚠️ No parent elements found but ignoring: {parent_selector}"
            )
            return
        else:
            raise RuntimeError(
                f"No parent elements found for selector: {parent_selector} "
                f"{'with text: ' + filter_text if filter_text else ''}"
            )
    logger.info(f"🔍 Found {total} parent element(s) for: {parent_selector}")

    # Select which parents to process
    parent_indices: List[int]
    if parent_idx is not None:
        if parent_idx < 0 or parent_idx >= total:
            if ignore_error:
                logger.warning(
                    f"⚠️ array_select_one index {parent_idx} is out of range (found {total}), but ignoring error."
                )
                return
            else:
                raise RuntimeError(
                    f"array_select_one index {parent_idx} is out of range (found {total})."
                )
        parent_indices = [parent_idx]
    else:
        parent_indices = list(range(total))

    clicks: List[Dict[str, Any]] = get_key(step, "click", default=[])
    if not isinstance(clicks, list) or not clicks:
        raise RuntimeError('Missing non-empty "click" array for array step.')

    # For each selected parent, run the child clicks in order
    for i in parent_indices:
        p = parents.nth(i)
        logger.info(f"🔄 Processing parent index {i}...")
        for j, child in enumerate(clicks, start=1):
            ctag = get_key(child, "tag")
            ctext = get_key(child, "text")
            cattr = get_key(child, "attr", "arrt", "attribute")
            cvalue = get_key(child, "value")
            ccls = get_key(child, "class")
            csleep = get_key(child, "sleep")
            cignore = get_key(child, "ignore", default=False)

            child_selector = build_css_selector(ctag, ccls, cattr, cvalue)
            child_loc = p.locator(child_selector)
            if ctext:
                child_loc = child_loc.filter(has_text=ctext)

            logger.info(
                f"  🔘 Child click [{j}]: {child_selector}{' | has_text=' + ctext if ctext else ''}"
            )
            try:
                success = wait_and_click(
                    child_loc,
                    index=0,
                    timeout=float(get_key(step, "timeout", default=35000)),
                    ignore_error=cignore,
                )
                if not success and cignore:
                    continue
            except PWTimeout as e:
                if cignore:
                    logger.warning(
                        f"⚠️ Timeout waiting for child element but ignoring: {child_selector}"
                    )
                    continue
                else:
                    raise RuntimeError(
                        f"Timeout waiting for child element: {child_selector}"
                    ) from e
            step_sleep(csleep)

    step_sleep(get_key(step, "sleep"))


def exec_step_frame(page, step: Dict[str, Any]):
    """Switch to an iframe."""
    return switch_to_frame(page, step)


def exec_step_refresh(page, step: Dict[str, Any]) -> None:
    """Reload the current page."""
    logger.info("🔄 Refreshing the current page")
    page.reload()
    step_sleep(get_key(step, "sleep"))


def exec_step_main_frame(page, step: Dict[str, Any]):
    """Switch back to the main frame."""
    switch_to_main_frame(page)
    step_sleep(get_key(step, "sleep"))
    return None


def exec_step_use_last_tab(browser, step: Dict[str, Any]):
    """Switch to the last opened tab."""
    tabs = browser.pages
    if len(tabs) > 1:
        last_tab = tabs[-1]
        last_tab.bring_to_front()
        logger.info(f"📑 Switched to last tab: {last_tab.url}")
    else:
        logger.info("ℹ️ Only one tab open, no switch needed.")
    step_sleep(get_key(step, "sleep"))


def exec_step_scroll(
    page, step: Dict[str, Any], current_frame=None, parent=None
) -> None:
    """Scroll to an element or by position."""
    tag = get_key(step, "tag")
    attr = get_key(step, "attr", "arrt", "attribute")
    value = get_key(step, "value")
    cls = get_key(step, "class")
    text = get_key(step, "text")
    idx = to_int_or_none(get_key(step, "array_select_one"))
    ignore_error = get_key(step, "ignore", default=False)

    # Check if it's a position scroll
    x = get_key(step, "x")
    y = get_key(step, "y")

    if x is not None or y is not None:
        # Position-based scrolling
        x_pos = int(x) if x is not None else 0
        y_pos = int(y) if y is not None else 0
        logger.info(f"📜 Scrolling to position: x={x_pos}, y={y_pos}")
        page.evaluate(f"window.scrollTo({x_pos}, {y_pos})")
        return

    # Element-based scrolling
    if not any([tag, attr, value, cls, text]):
        raise RuntimeError(
            "Scroll step requires either position (x,y) or element selector"
        )

    selector = build_css_selector(tag, cls, attr, value)

    root = get_locator_root(page, current_frame, parent)
    loc = root.locator(selector)

    if text:
        loc = loc.filter(has_text=text)

    logger.info(f"📜 Scroll to selector: {selector}")

    try:
        if idx is None:
            idx = 0

        count = loc.count()
        if count == 0:
            if ignore_error:
                logger.warning(
                    f"⚠️ No elements found for scrolling but ignoring: {selector}"
                )
                return
            else:
                raise RuntimeError(f"No elements found for scrolling: {selector}")

        if idx < 0 or idx >= count:
            if ignore_error:
                logger.warning(
                    f"⚠️ array_select_one index {idx} is out of range (found {count}), but ignoring error."
                )
                return
            else:
                raise RuntimeError(
                    f"array_select_one index {idx} is out of range (found {count})."
                )

        target = loc.nth(idx)
        target.wait_for(
            state="visible", timeout=float(get_key(step, "timeout", default=35000))
        )
        target.scroll_into_view_if_needed()
        logger.info("✅ Scrolled to element successfully")

    except Exception as e:
        if ignore_error:
            logger.warning(f"⚠️ Scroll failed but ignoring: {e}")
        else:
            raise


def download_requests(url, out_path, retries=3):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/123.0.6312.86 Safari/537.36",
        "Accept": "*/*",
        "Accept-Language": "en-US,en;q=0.9",
        "Range": "bytes=0-",
        "Referer": "",
    }
    session = requests.Session()
    for attempt in range(1, retries + 1):
        try:
            with session.get(url, headers=headers, stream=True, timeout=60) as r:
                print(
                    "HTTP",
                    r.status_code,
                    r.headers.get("Content-Length"),
                    r.headers.get("Accept-Ranges"),
                )
                # قبول کد وضعیت‌های 200, 202, 206 به عنوان موفقیت‌آمیز
                if r.status_code not in (200, 202, 206):
                    print(f"Not OK status, retrying... {r.status_code}")
                    time.sleep(1)
                    continue
                total = r.headers.get("Content-Length")
                with open(out_path, "wb") as f:
                    for chunk in r.iter_content(chunk_size=1024 * 16):
                        if chunk:
                            f.write(chunk)
                print("Saved to", out_path)
                return True
        except Exception as e:
            print(f"Attempt {attempt} failed:", e)
            time.sleep(1)
    return False


# در ابتدای فایل، کتابخانه‌های مورد نیاز را اضافه می‌کنم
import html
import os
import re


# تابع جدید برای استخراج محتوای VTT از HTML
def extract_vtt_content(html_content):
    """
    استخراج محتوای واقعی VTT از HTML دریافت شده
    """
    # روش اول: استخراج محتوای داخل تگ <pre>
    pre_match = re.search(
        r"<pre[^>]*>(.*?)</pre>", html_content, re.DOTALL | re.IGNORECASE
    )
    if pre_match:
        content = pre_match.group(1)
        # حذف تگ‌های HTML اضافی
        content = re.sub(r"<[^>]+>", "", content)
        # حذف کاراکترهای HTML entity
        content = html.unescape(content)
        # حذف فاصله‌های اضافی در ابتدا و انتها
        content = content.strip()
        logger.info("✅ محتوای VTT از تگ <pre> استخراج شد")
        return content

    # روش دوم: اگر تگ <pre> وجود نداشته باشد، کل بدنه را بررسی کن
    body_match = re.search(
        r"<body[^>]*>(.*?)</body>", html_content, re.DOTALL | re.IGNORECASE
    )
    if body_match:
        content = body_match.group(1)
        content = re.sub(r"<[^>]+>", "", content)
        content = html.unescape(content)
        content = content.strip()
        logger.info("⚠️ محتوای VTT از بدنه صفحه استخراج شد (بدون تگ <pre>)")
        return content

    # روش سوم: اگر هیچکدام کار نکرد، کل محتوا را برگردان
    logger.warning("⚠️ نتوانستم محتوای VTT را استخراج کنم، کل محتوا استفاده می‌شود")
    return html_content


# تابع جدید برای دانلود مستقیم زیرنویس‌ها با Playwright
def download_subtitle_direct(url, output_path, page_context):
    """
    دانلود مستقیم فایل زیرنویس با Playwright و پردازش محتوا
    """
    logger.info(f"🎬 در حال دانلود فایل زیرنویس از: {url}")

    try:
        # استفاده از context موجود برای ایجاد صفحه جدید
        new_page = page_context.new_page()

        # افزودن اسکریپت برای جلوگیری از تشخیص اتوماسیون
        new_page.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', { get: () => false });
        window.navigator.chrome = { runtime: {}, app: {} };
        Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en'] });
        Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
        """)

        # باز کردن URL
        logger.info("در حال بارگذاری صفحه...")
        response = new_page.goto(url, wait_until="networkidle", timeout=60000)
        new_page.evaluate("window.scrollTo(0, document.body.scrollHeight);")
        if not response:
            logger.error("❌ خطا در بارگذاری صفحه: پاسخ دریافت نشد")
            new_page.close()
            return False

        logger.info(f"کد وضعیت HTTP: {response.status}")

        # اگر کد وضعیت 202 باشد، صبر کنیم تا محتوا بارگذاری شود
        if response.status == 202:
            logger.info(
                "⏳ دریافت کد وضعیت 202 (Accepted)، در حال انتظار برای محتوا..."
            )
            # حداکثر 10 ثانیه صبر کن
            for i in range(10):
                page_content = new_page.content()
                if "WEBVTT" in page_content or "<pre>" in page_content.lower():
                    logger.info(f"✅ محتوا پس از {i + 1} ثانیه بارگذاری شد")
                    break
                time.sleep(1)

        # دریافت محتوای کامل صفحه
        html_content = new_page.content()
        logger.info(f"📄 محتوای صفحه دریافت شد (طول: {len(html_content)} کاراکتر)")

        # استخراج محتوای VTT از HTML
        vtt_content = extract_vtt_content(html_content)
        logger.info(f"📝 محتوای استخراج شده VTT (طول: {len(vtt_content)} کاراکتر)")

        # بررسی محتوای استخراج شده
        if not vtt_content or len(vtt_content) < 10:
            logger.error("❌ محتوای استخراج شده خالی یا بسیار کوتاه است")
            # ذخیره محتوای HTML برای دیباگ
            debug_path = output_path + ".debug.html"
            with open(debug_path, "w", encoding="utf-8") as f:
                f.write(html_content)
            logger.info(f"🔍 محتوای دیباگ در {debug_path} ذخیره شد")
            new_page.close()
            return False

        # ذخیره محتوای VTT در فایل
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(vtt_content)

        logger.info(f"✅ فایل زیرنویس با موفقیت ذخیره شد در: {output_path}")
        logger.info(f"📊 اندازه فایل: {os.path.getsize(output_path)} بایت")

        new_page.close()
        return True

    except Exception as e:
        logger.error(f"❌ خطای غیرمنتظره: {str(e)}")
        return False


# تابع exec_step_download_from_link را به‌روزرسانی می‌کنم
def exec_step_download_from_link(
    page, step: Dict[str, Any], current_frame=None, parent=None
) -> None:
    """
    Click a link and save the downloaded file with specified extension
    Supports custom file extensions like vtt, mp4, pdf, etc.
    """
    tag = get_key(step, "tag")
    attr = get_key(step, "attr", "arrt", "attribute")
    value = get_key(step, "value")
    cls = get_key(step, "class")
    text = get_key(step, "text")
    idx = to_int_or_none(get_key(step, "array_select_one"))
    ignore_error = get_key(step, "ignore", default=False)
    download_dir = get_key(step, "download_dir", "dir", default=os.getcwd())
    filename = get_key(step, "filename", "file_name", "file")
    # دریافت پسوند فایل از workflow یا تشخیص خودکار
    file_extension = get_key(step, "extension", "file_extension", "ext")
    timeout = float(get_key(step, "timeout", default=70000))
    index = get_key(step, "index", default=1)
    selector = build_css_selector(tag, cls, attr, value)
    root = get_locator_root(page, current_frame, parent)
    loc = root.locator(selector)
    if text:
        loc = loc.filter(has_text=text)
    logger.info(
        f"📥 Download-from-link selector: {selector}{' | has_text=' + text if text else ''}"
    )
    try:
        if idx is None:
            idx = 0
        count = loc.count()
        if count == 0:
            if ignore_error:
                logger.warning(
                    f"⚠️ No elements found for download_from_link but ignoring: {selector}"
                )
                return
            else:
                raise RuntimeError(
                    f"No elements found for download_from_link: {selector}"
                )
        if idx < 0 or idx >= count:
            if ignore_error:
                logger.warning(
                    f"⚠️ array_select_one index {idx} is out of range (found {count}), but ignoring error."
                )
                return
            else:
                raise RuntimeError(
                    f"array_select_one index {idx} is out of range (found {count})."
                )
        target = loc.nth(idx)
        target.wait_for(
            state="visible",
            timeout=timeout,
        )
        target.scroll_into_view_if_needed()
        # Get the href attribute which contains the download link
        download_url = target.get_attribute("href")
        if not download_url:
            raise RuntimeError("No download link (href) found in the target element.")
        # Convert relative URLs to absolute
        if not download_url.startswith(("http://", "https://")):
            base_url = page.url
            download_url = urljoin(base_url, download_url)
            logger.info(f"🔄 Converted relative URL to absolute: {download_url}")
        logger.info(f"📥 Found download link: {download_url}")
        # استخراج پسوند فایل از URL اگر در workflow مشخص نشده باشد
        if not file_extension:
            parsed_url = urlparse(download_url)
            query_params = parsed_url.query.split("&")
            for param in query_params:
                if param.startswith("fileExtension="):
                    file_extension = param.split("=")[1]
                    break
            # اگر از URL استخراج نشد، از آخرین بخش مسیر URL استفاده کن
            if not file_extension:
                path = parsed_url.path
                if "." in path:
                    file_extension = path.split(".")[-1]
        # پاک کردن کاراکترهای غیرمجاز از پسوند
        if file_extension:
            file_extension = re.sub(r'[\\/*?:"<>|]', "", file_extension).lower()
            # اگر پسوند با نقطه شروع نشده، نقطه اضافه کن
            if not file_extension.startswith("."):
                file_extension = f".{file_extension}"
        else:
            file_extension = ".mp4"  # پیش‌فرض

        # بررسی اینکه آیا پسوند مربوط به زیرنویس است
        is_subtitle = False
        subtitle_extensions = ["vtt", "str"]
        clean_extension = file_extension.strip().lower()
        if clean_extension.startswith("."):
            clean_extension = clean_extension[1:]
        if clean_extension in subtitle_extensions:
            is_subtitle = True

        # ایجاد نام فایل با پسوند مناسب
        page_title = page.title() or "download"
        safe_title = make_safe_filename(page_title, default="download", ext="")
        out_path = os.path.join(download_dir, f"{safe_title}_{index}{file_extension}")
        # ایجاد دایرکتوری اگر وجود نداشته باشد
        os.makedirs(download_dir, exist_ok=True)

        if is_subtitle:
            # استفاده از روش ویژه برای دانلود زیرنویس
            logger.info(f"🎬 در حال دانلود زیرنویس با پسوند {clean_extension}...")
            success = download_subtitle_direct(download_url, out_path, page.context)
            if success:
                logger.info(f"✅ زیرنویس با موفقیت دانلود شد: {out_path}")
            else:
                logger.warning(f"⚠️ دانلود زیرنویس با شکست مواجه شد.")
                # به عنوان پشتیبان، سعی در دانلود مستقیم
                logger.info("🔄 تلاش برای دانلود مستقیم به عنوان روش پشتیبان...")
                success = download_requests(download_url, out_path)
                if success:
                    logger.info(f"✅ دانلود مستقیم موفقیت‌آمیز بود: {out_path}")
                else:
                    logger.error("❌ هر دو روش دانلود شکست خوردند.")
        else:
            # استفاده از روش معمول برای دانلود فایل‌های دیگر (ویدیو، pdf و...)
            success = download_requests(download_url, out_path)
            if success:
                logger.info(f"💾 File downloaded successfully: {out_path}")
            else:
                logger.warning(f"⚠️ File download failed.")
    except Exception as e:
        if ignore_error:
            logger.warning(f"⚠️ download_from_link failed but ignoring: {e}")
        else:
            raise
    step_sleep(get_key(step, "sleep"))


# ------------------ group_action ------------------
def exec_step_group_action(
    page, browser, step: Dict[str, Any], current_frame=None, parent=None
) -> None:
    """
    group_action:
    - find multiple elements (parents)
    - run nested actions on each parent
    Supports:
      - "global_actions": true  -> run actions against page (global) instead of parent
      - action-level "global": true -> that single action runs against page
    """
    tag = get_key(step, "tag")
    attr = get_key(step, "attr", "arrt", "attribute")
    value = get_key(step, "value")
    cls = get_key(step, "class")
    filter_text = get_key(step, "if_find_text_inside")
    parent_idx = to_int_or_none(get_key(step, "array_select_one"))  # optional
    ignore_error = get_key(step, "ignore", default=False)
    timeout = float(get_key(step, "timeout", default=35000))
    # NEW: group-level global flag
    group_global = bool(get_key(step, "global_actions", default=False))

    parent_selector = build_css_selector(tag, cls, attr, value)

    root = get_locator_root(page, current_frame, parent)
    parents = root.locator(parent_selector)

    if filter_text:
        parents = parents.filter(has_text=filter_text)

    total = parents.count()
    if total == 0:
        if ignore_error:
            logger.warning(
                f"⚠️ No parent elements found for group_action but ignoring: {parent_selector}"
            )
            return
        else:
            raise RuntimeError(
                f"No parent elements found for group_action selector: {parent_selector} "
                f"{'with text: ' + filter_text if filter_text else ''}"
            )

    logger.info(
        f"🧩 group_action: found {total} parent element(s) for: {parent_selector}"
    )

    # select parent indices
    if parent_idx is not None:
        if parent_idx < 0 or parent_idx >= total:
            if ignore_error:
                logger.warning(
                    f"⚠️ group_action array_select_one index {parent_idx} is out of range (found {total}), but ignoring."
                )
                return
            else:
                raise RuntimeError(
                    f"group_action array_select_one index {parent_idx} is out of range (found {total})."
                )
        parent_indices = [parent_idx]
    else:
        parent_indices = list(range(total))

    actions: List[Dict[str, Any]] = get_key(step, "actions", "steps", default=[])
    if not isinstance(actions, list) or not actions:
        raise RuntimeError('group_action requires non-empty "actions" array.')

    # for each selected parent, run actions
    for i in parent_indices:
        p = parents.nth(i)
        logger.info(f"🧩 [group_action] Processing parent index {i}...")
        try:
            p.wait_for(state="visible", timeout=timeout)
        except Exception:
            pass
        try:
            p.scroll_into_view_if_needed()
        except Exception:
            pass

        local_frame = current_frame

        for j, action in enumerate(actions, start=1):
            a_title = get_key(
                action, "title", "Title", default=f"group_action action #{j}"
            )
            a_type = get_key(action, "type")
            if not a_type:
                logger.warning(
                    "⚠️ [group_action] Missing 'type' in nested action, skipping."
                )
                continue

            stype_l = str(a_type).strip().lower()
            logger.info(
                f"   ▶️ [group_action] Parent {i} - Action {j}: {a_title} ({stype_l})"
            )

            action_ignore = get_key(action, "ignore", default=False)
            # action-level global (per-action)
            action_global = bool(get_key(action, "global", default=False))

            # Decide effective parent for this action:
            # - If group_global True => actions act on page (parent=None)
            # - Else if action_global True => action acts on page (parent=None)
            # - Else => action acts inside current parent 'p' (parent=p)
            effective_parent = None if (group_global or action_global) else p

            try:
                if stype_l == "click":
                    exec_step_click(page, action, local_frame, parent=effective_parent)
                elif stype_l == "write":
                    exec_step_write(page, action, local_frame, parent=effective_parent)
                elif stype_l == "scroll":
                    exec_step_scroll(page, action, local_frame, parent=effective_parent)
                elif stype_l == "array":
                    exec_step_array(page, action, local_frame, parent=effective_parent)
                elif stype_l == "group_action":
                    exec_step_group_action(
                        page, browser, action, local_frame, parent=effective_parent
                    )
                elif stype_l == "download_from_link":
                    exec_step_download_from_link(
                        page, action, local_frame, parent=effective_parent
                    )
                # elif stype_l in ("download_page", "save_page"):
                #     exec_step_download_page(page, action)
                elif stype_l == "use_last_tab":
                    exec_step_use_last_tab(browser, action)
                elif stype_l == "goto":
                    exec_step_goto(page, action)
                    local_frame = None
                elif stype_l == "frame":
                    local_frame = exec_step_frame(page, action)
                elif stype_l == "main_frame":
                    local_frame = exec_step_main_frame(page, action)
                else:
                    if action_ignore or ignore_error:
                        logger.warning(
                            f"⚠️ [group_action] Unsupported nested action type but ignoring: '{a_type}'"
                        )
                    else:
                        raise RuntimeError(
                            f"[group_action] Unsupported nested action type: '{a_type}'"
                        )

            except Exception as e:
                if action_ignore or ignore_error:
                    logger.warning(
                        f"⚠️ [group_action] Nested action failed but ignoring: {a_title} | {e}"
                    )
                    continue
                else:
                    raise

    step_sleep(get_key(step, "sleep"))


# ------------------ Runner ------------------
def run(
    workflow: List[Dict[str, Any]],
    start_url: Optional[str] = None,
    profile_dir: Optional[str] = None,
):
    width, height = 1300, 900
    profile = profile_dir or os.path.join(os.getcwd(), "pw_profile")

    logger.info("🚀 === Starting workflow run ===")
    logger.info(f"📁 Profile dir: {profile}")
    logger.info(f"🖥️ Viewport: {width}x{height}")

    CHROME_UA = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/123.0.6312.86 Safari/537.36"
    )
    LOCALE = "en-US"
    ACCEPT_LANG = "en-US,en;q=0.9"
    TIMEZONE_ID = "Asia/Tehran"

    chromium_args = [
        f"--window-size={width},{height}",
        "--start-maximized",
        "--disable-blink-features=AutomationControlled",
        "--disable-infobars",
        "--no-default-browser-check",
        "--no-first-run",
        "--disable-features=IsolateOrigins,site-per-process",
    ]

    stealth_js = r"""
(() => {
  try {
    Object.defineProperty(navigator, 'webdriver', { get: () => false, configurable: true });
    try { Object.defineProperty(navigator, 'languages', { get: () => ['en-US','en'], configurable: true }); } catch (e) {}
    try { Object.defineProperty(navigator, 'plugins', { get: () => [{name:'Chrome PDF Plugin', filename:'internal-pdf-viewer'}], configurable: true }); } catch (e) {}
    try { Object.defineProperty(navigator, 'mimeTypes', { get: () => [{type:'application/pdf', suffixes:'pdf'}], configurable: true }); } catch (e) {}
    try { window.chrome = window.chrome || { runtime: {} }; } catch (e) {}
    try {
      Object.defineProperty(navigator, 'platform', { get: () => 'Win32', configurable: true });
      Object.defineProperty(navigator, 'vendor', { get: () => 'Google Inc.', configurable: true });
      Object.defineProperty(navigator, 'appVersion', { get: () => '5.0 (Windows)', configurable: true });
    } catch (e) {}
    try { Object.defineProperty(navigator, 'hardwareConcurrency', { get: () => 8, configurable: true }); } catch (e) {}
    try { Object.defineProperty(navigator, 'deviceMemory', { get: () => 8, configurable: true }); } catch (e) {}
    try {
      const origQuery = navigator.permissions && navigator.permissions.query;
      if (origQuery) {
        navigator.permissions.query = function(parameters) {
          if (parameters && parameters.name === 'notifications') {
            return Promise.resolve({ state: Notification.permission });
          }
          return origQuery(parameters);
        };
      }
    } catch (e) {}
    try {
      if (navigator.userAgentData && navigator.userAgentData.brands) {
        navigator.userAgentData.brands = [{brand: "Chromium", version: "123"}, {brand: "Google Chrome", version: "123"}];
      }
    } catch (e) {}
    try {
      const getParameter = WebGLRenderingContext.prototype.getParameter;
      WebGLRenderingContext.prototype.getParameter = function(parameter) {
        if (parameter === 37445) return "Intel Inc.";
        if (parameter === 37446) return "Intel(R) HD Graphics 620";
        return getParameter.call(this, parameter);
      };
    } catch (e) {}
    try {
      const originalToString = Function.prototype.toString;
      const myToString = function() {
        if (this === navigator.permissions.query) {
          return 'function query() { [native code] }';
        }
        return originalToString.apply(this, arguments);
      };
      Function.prototype.toString = myToString;
    } catch (e) {}
  } catch (err) {}
})();
"""

    fatal_error: Optional[Exception] = None

    with sync_playwright() as p:
        browser = p.chromium.launch_persistent_context(
            user_data_dir=profile,
            headless=False,
            args=chromium_args,
            viewport={"width": width, "height": height},
            screen={"width": width, "height": height},
            accept_downloads=True,
            user_agent=CHROME_UA,
            locale=LOCALE,
            timezone_id=TIMEZONE_ID,
            extra_http_headers={"Accept-Language": ACCEPT_LANG},
        )

        try:
            browser.add_init_script(stealth_js)
            logger.info("🔐 Stealth init script injected.")
        except Exception as e:
            logger.warning(f"⚠️ Failed to add stealth init script: {e}")

        page = browser.pages[0] if browser.pages else browser.new_page()
        current_frame = None

        # Optional initial URL
        if start_url:
            logger.info(f"🌐 Initial goto: {start_url}")
            page.goto(start_url)

        # Execute steps (do NOT close browser on failures)
        for idx, step in enumerate(workflow, start=1):
            title = get_key(step, "title", "Title", default=f"Step #{idx}")
            stype = get_key(step, "type")
            ignore_error = get_key(step, "ignore", default=False)

            logger.info(f"--- Step {idx}: {title} ---")
            print(f"📝 [Step {idx}] {title}")

            if not stype:
                if ignore_error:
                    logger.warning("⚠️ Missing 'type' in step, but ignoring error.")
                    continue
                fatal_error = RuntimeError('Missing "type" in step.')
                logger.error(f"❌ Step failed: {title} | {fatal_error}")
                print(f"❌ [ERROR] {title}: {fatal_error}")
                break

            stype_l = str(stype).strip().lower()

            try:
                if stype_l == "goto":
                    exec_step_goto(page, step)
                    current_frame = None
                elif stype_l == "click":
                    exec_step_click(page, step, current_frame)
                elif stype_l == "select":
                    exec_step_select(page, step, current_frame)
                elif stype_l == "group_excel":
                    exec_step_group_excel(page, browser, step, current_frame)
                elif stype_l == "array":
                    exec_step_array(page, step, current_frame)
                elif stype_l == "refresh":
                    exec_step_refresh(page, step)
                elif stype_l == "group_action":
                    exec_step_group_action(page, browser, step, current_frame)
                elif stype_l == "frame":
                    current_frame = exec_step_frame(page, step)
                elif stype_l == "main_frame":
                    current_frame = exec_step_main_frame(page, step)
                elif stype_l == "write":
                    exec_step_write(page, step, current_frame)
                elif stype_l == "use_last_tab":
                    exec_step_use_last_tab(browser, step)
                elif stype_l == "scroll":
                    exec_step_scroll(page, step, current_frame)
                elif stype_l == "download_from_link":
                    exec_step_download_from_link(page, step, current_frame)
                else:
                    if ignore_error:
                        logger.warning(f"⚠️ Unsupported step type but ignoring: '{stype}'")
                    else:
                        raise RuntimeError(f'Unsupported step type: "{stype}"')
            except Exception as e:
                if ignore_error:
                    logger.warning(f"⚠️ Step failed but ignoring: {title} | {e}")
                    print(f"⚠️ [WARNING] {title}: {e}")
                    continue

                # Stop the automation, but keep the browser open for manual inspection
                fatal_error = e
                logger.error(f"❌ Step failed: {title} | {e}")
                print(f"❌ [ERROR] {title}: {e}")
                break

        if fatal_error is None:
            logger.info("✅ === Workflow completed successfully ===")
            print("✅ Workflow completed successfully.")
        else:
            logger.warning("🛑 === Workflow stopped due to an error (browser stays open) ===")
            print("🛑 Workflow stopped due to an error (browser stays open).")

        # Keep the browser open until the user closes it manually
        logger.info("🧭 Close the browser window to finish the script (no auto-close).")
        print("🧭 Close the browser window to finish the script (no auto-close).")

        while True:
            try:
                # context.pages returns currently open pages
                if len(browser.pages) == 0:
                    break
            except Exception:
                break
            time.sleep(0.5)

        # After the user closes the browser, propagate the error (if any)
        if fatal_error is not None:
            raise fatal_error



def run_course_automation(workflow_path):
    print("Run run_course_automation")
    global LOG_CAPTURE_LIST
    LOG_CAPTURE_LIST.clear()

    try:
        if not os.path.exists(workflow_path):
            raise FileNotFoundError("Workflow file missing")

        with open(workflow_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # run() now keeps browser open until user closes it, even on errors
        run(data)

        return True, "\n".join(LOG_CAPTURE_LIST)

    except Exception as e:
        return False, "\n".join(LOG_CAPTURE_LIST) + f"\nFATAL: {e}"

